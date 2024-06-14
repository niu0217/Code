# 三方库
from openpyxl import load_workbook
from openpyxl import Workbook


def modifyName(filename):
    wb = load_workbook(filename)
    ws = wb.active

    for row in range(2, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            if col == 1:
                ws.cell(row=row, column=col, value="niu0217")

    wb.save(filename)


if __name__ == "__main__":
    modifyName('chinaData.xlsx')
