# 三方库
from openpyxl import load_workbook
from openpyxl import Workbook


def modifyName(filename, outfilename):
    wb = load_workbook(filename)
    ws = wb.active
    for row in range(2, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            if col == 1:
                ws.cell(row=row, column=col, value="bœuf")
    wb_new = Workbook()
    ws_new = wb_new.active
    for row in ws.iter_rows(values_only=True):
        ws_new.append(row)
    wb_new.save(outfilename)


if __name__ == "__main__":
    modifyName('frData.xlsx', 'output.xlsx')
